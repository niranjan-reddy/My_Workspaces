import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["inventory"]

products_per_supplier = {}
total_value_per_supplier = {}
products_under_55_inv = {}

inventory_price = product_list.cell(1, 5).value = "Total Cost of Inventory"

# for loop to iterate through the items
# for product_row in product_list.max_row:
# we need to start iterating from row 2 and not 1 hence we need to follow the following. Also we need to provide the last item in the range as Python considers the last item's index which is EXCLUSIVE. so we must increase it by 1.
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    #print(supplier_name)


    # calculation of number of products per supplier
    if supplier_name in products_per_supplier:
        #current_num_products = products_per_supplier[supplier_name]
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        #print("adding new supplier")
        products_per_supplier[supplier_name] = 1

    # calculation of total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # logic products with inventory with less than 55
    if inventory < 55:
        products_under_55_inv[product_num] = inventory

    # logic to add the value for total inventory price
    inventory_price.value = inventory * price

inv_file.save("inventory_with_total_value.xlsx")

print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_55_inv)
